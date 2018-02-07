from openpyxl import Workbook

from flask import Blueprint, Response, send_from_directory
from flask_restful import Api
from flasgger import swag_from

from app.docs.admin.apply.goingout import *
from app.models.account import StudentModel
from app.views import BaseResource, admin_only

from utils.excel_style_manager import get_cell_positions_by_student_number, ready_applyment_worksheet

api = Api(Blueprint('admin-goingout-api', __name__))
api.prefix = '/admin'


@api.resource('/goingout')
class GoingoutDownload(BaseResource):
    @swag_from(GOINGOUT_DOWNLOAD_GET)
    @admin_only
    def get(self):
        """
        외출신청 엑셀 다운로드
        """
        wb = Workbook()
        ws = wb.active

        ready_applyment_worksheet(ws)

        for student in StudentModel.objects:
            number_cell, name_cell, status_cell = get_cell_positions_by_student_number(student.number)

            ws[number_cell] = student.number
            ws[name_cell] = student.name

            goingout_apply = student.goingout_apply

            if goingout_apply.on_saturday and goingout_apply.on_sunday:
                status = '토요일, 일요일 외출'
            elif goingout_apply.on_saturday:
                status = '토요일 외출'
            elif goingout_apply.on_sunday:
                status = '일요일 외출'
            else:
                continue

            ws[status_cell] = status

        wb.save('goingout.xlsx')
        wb.close()

        response = Response(send_from_directory('../', 'goingout.xlsx', as_attachment=True))
        response.headers['Content-Type'] = 'application/vnd.ms-excel'
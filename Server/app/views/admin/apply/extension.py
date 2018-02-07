from openpyxl import Workbook

from flask import Blueprint, Response, send_from_directory
from flask_restful import Api
from flasgger import swag_from

from app.docs.admin.apply.extension import *
from app.models.account import StudentModel
from app.views import BaseResource, admin_only

from utils.excel_style_manager import get_cell_positions_by_student_number, ready_applyment_worksheet

EXTENSION_CLASSES = ['', '가온실', '나온실', '다온실', '라온실', '3층 독서실', '4층 독서실', '5층 열린교실']

api = Api(Blueprint('admin-extension-api', __name__))
api.prefix = '/admin/extension'


@api.resource('/11')
class Extension11Download(BaseResource):
    @swag_from(EXTENSION_DOWNLOAD_GET)
    @admin_only
    def get(self):
        """
        11시 연장신청 엑셀 다운로드
        """
        wb = Workbook()
        ws = wb.active

        ready_applyment_worksheet(ws)

        for student in StudentModel.objects:
            number_cell, name_cell, status_cell = get_cell_positions_by_student_number(student.number)

            ws[number_cell] = student.number
            ws[name_cell] = student.name

            extension_apply = student.extension_apply_11

            if not extension_apply:
                continue

            ws[status_cell] = EXTENSION_CLASSES[extension_apply.class_]

        wb.save('11.xlsx')
        wb.close()

        response = Response(send_from_directory('../', '11.xlsx', as_attachment=True))
        response.headers['Content-Type'] = 'application/vnd.ms-excel'

        return response


@api.resource('/12')
class Extension12Download(BaseResource):
    @swag_from(EXTENSION_DOWNLOAD_GET)
    @admin_only
    def get(self):
        """
        12시 연장신청 엑셀 다운로드
        """
        wb = Workbook()
        ws = wb.active

        ready_applyment_worksheet(ws)

        for student in StudentModel.objects:
            number_cell, name_cell, status_cell = get_cell_positions_by_student_number(student.number)

            ws[number_cell] = student.number
            ws[name_cell] = student.name

            extension_apply = student.extension_apply_11

            if not extension_apply:
                continue

            ws[status_cell] = EXTENSION_CLASSES[extension_apply.class_]

        wb.save('12.xlsx')
        wb.close()

        response = Response(send_from_directory('../', '12.xlsx', as_attachment=True))
        response.headers['Content-Type'] = 'application/vnd.ms-excel'